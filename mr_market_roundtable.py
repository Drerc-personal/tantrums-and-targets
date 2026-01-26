#!/usr/bin/env python3
"""
===============================================================================
MR. MARKET ROUNDTABLE - Combined Daily Script
===============================================================================
Purpose: One script to rule them all. Run after market close to:
    1. Fetch market data for all 25 stocks
    2. Update tracker (positions, pending orders, daily snapshot)
    3. Detect Track 2/3 alert candidates
    4. Generate a single AI Roundtable prompt file
    5. Save prompt to daily file for easy attachment to AI chats

Usage:
    python mr_market_roundtable.py

Output:
    - Updates mr_market_tracker.xlsx
    - Creates prompts/YYYY_MM_DD_roundtable_prompt.txt

Author:  Built with Claude, ChatGPT, and Gemini
Version: 1.0
Date:    January 2026
===============================================================================
"""

import yfinance as yf
from openpyxl import load_workbook
from datetime import datetime, timedelta
import os
import json
import argparse
import re

# =============================================================================
# BLOCK 1: CONFIGURATION
# =============================================================================
print("=" * 70)
print("MR. MARKET ROUNDTABLE - Combined Daily Script v1.0")
print(f"Run time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print("=" * 70)

# 1.1 - File paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TRACKER_FILE = os.path.join(SCRIPT_DIR, "mr_market_tracker.xlsx")
TRACK2_HISTORY_FILE = os.path.join(SCRIPT_DIR, "track2_trigger_history.json")
PROMPTS_DIR = os.path.join(SCRIPT_DIR, "prompts")

# 1.2 - Create directories if they don't exist
if not os.path.exists(PROMPTS_DIR):
    os.makedirs(PROMPTS_DIR)
    print(f"    Created prompts directory: {PROMPTS_DIR}")

DECISIONS_DIR = os.path.join(SCRIPT_DIR, "decisions")
if not os.path.exists(DECISIONS_DIR):
    os.makedirs(DECISIONS_DIR)
    print(f"    Created decisions directory: {DECISIONS_DIR}")

# 1.3 - All 25 tickers
TICKERS = [
    "VOO", "MSFT", "AAPL", "NVDA", "ASML", "BAC", "CB", "V", "MCO",
    "JNJ", "NVO", "CVX", "GEV", "GE", "WM",
    "ODFL", "TYL", "FICO", "CPRT", "IDXX", "VRSN", "ROP", "RSG", "JKHY", "PWR"
]

# 1.4 - Watchlist with company names and strategy tags
WATCHLIST = {
    "VOO":  {"name": "S&P 500 Index", "strategy": "CORE", "tier": "1"},
    "MSFT": {"name": "Microsoft", "strategy": "CORE", "tier": "1"},
    "AAPL": {"name": "Apple", "strategy": "CORE", "tier": "1"},
    "NVDA": {"name": "NVIDIA", "strategy": "CORE", "tier": "1"},
    "ASML": {"name": "ASML", "strategy": "CORE", "tier": "1"},
    "BAC":  {"name": "Bank of America", "strategy": "CORE", "tier": "1"},
    "CB":   {"name": "Chubb", "strategy": "CORE", "tier": "1"},
    "V":    {"name": "Visa", "strategy": "CORE", "tier": "1"},
    "MCO":  {"name": "Moody's", "strategy": "CORE", "tier": "1"},
    "JNJ":  {"name": "Johnson & Johnson", "strategy": "CORE", "tier": "1"},
    "NVO":  {"name": "Novo Nordisk", "strategy": "CORE", "tier": "1"},
    "CVX":  {"name": "Chevron", "strategy": "CORE", "tier": "1"},
    "GEV":  {"name": "GE Vernova", "strategy": "CORE", "tier": "1"},
    "GE":   {"name": "GE Aerospace", "strategy": "CORE", "tier": "1"},
    "WM":   {"name": "Waste Management", "strategy": "CORE", "tier": "1"},
    "ODFL": {"name": "Old Dominion Freight", "strategy": "HUNT", "tier": "1B"},
    "TYL":  {"name": "Tyler Technologies", "strategy": "DCA", "tier": "1B"},
    "FICO": {"name": "Fair Isaac", "strategy": "DCA", "tier": "1B"},
    "CPRT": {"name": "Copart", "strategy": "HUNT", "tier": "1B"},
    "IDXX": {"name": "IDEXX Labs", "strategy": "DCA", "tier": "1B"},
    "VRSN": {"name": "Verisign", "strategy": "HUNT", "tier": "1B"},
    "ROP":  {"name": "Roper Technologies", "strategy": "HUNT", "tier": "1B"},
    "RSG":  {"name": "Republic Services", "strategy": "DCA", "tier": "1B"},
    "JKHY": {"name": "Jack Henry", "strategy": "DCA", "tier": "1B"},
    "PWR":  {"name": "Quanta Services", "strategy": "HUNT", "tier": "1B"},
}

# 1.5 - Tri-Anchor Target Prices (Track 3)
TARGETS = {
    "VOO":  {"target": 570, "add_target": 540},
    "MSFT": {"target": 450, "add_target": 445},
    "AAPL": {"target": 230, "add_target": 220},
    "NVDA": {"target": 150, "add_target": 140},
    "ASML": {"target": 900, "add_target": 850},
    "BAC":  {"target": 48, "add_target": 45},
    "CB":   {"target": 280, "add_target": 270},
    "V":    {"target": 310, "add_target": 300},
    "MCO":  {"target": 450, "add_target": 430},
    "JNJ":  {"target": 180, "add_target": 170},
    "NVO":  {"target": 50, "add_target": 45},
    "CVX":  {"target": 150, "add_target": 145},
    "GEV":  {"target": 550, "add_target": 500},
    "GE":   {"target": 280, "add_target": 270},
    "WM":   {"target": 210, "add_target": 200},
    "ODFL": {"target": 140, "add_target": 125},
    "TYL":  {"target": 400, "add_target": 380},
    "FICO": {"target": 1400, "add_target": 1300},
    "CPRT": {"target": 35, "add_target": 32},
    "IDXX": {"target": 550, "add_target": 500},
    "VRSN": {"target": 220, "add_target": 200},
    "ROP":  {"target": 400, "add_target": 380},
    "RSG":  {"target": 190, "add_target": 180},
    "JKHY": {"target": 160, "add_target": 150},
    "PWR":  {"target": 350, "add_target": 320},
}

# 1.6 - Exit Criteria (what would break the thesis)
EXIT_CRITERIA = {
    "VOO":  "N/A - index, always hold",
    "MSFT": "Azure growth <15% for 3+ quarters; loses enterprise cloud share to AWS/GCP",
    "AAPL": "iPhone unit decline >15% YoY for 2+ years; Services growth stalls <10%",
    "NVDA": "Loses GPU AI training dominance to AMD/custom silicon; data center growth <20%",
    "ASML": "EUV technology leapfrogged; China restrictions permanently impair >30% revenue",
    "BAC":  "Net interest margin compression <2% sustained; major credit losses in recession",
    "CB":   "Combined ratio >100% for 2+ years; catastrophic reserve deficiency",
    "V":    "Payment volume growth <5% sustained; regulatory cap on interchange fees",
    "MCO":  "Credit rating market share loss >10pts to S&P/Fitch; regulatory action on conflicts",
    "JNJ":  "Talc litigation exceeds $50B; pharma pipeline fails to offset LOEs",
    "NVO":  "GLP-1 competition erodes pricing power >30%; safety signal emerges",
    "CVX":  "Oil prices <$50 sustained 2+ years; fails energy transition pivot",
    "GEV":  "Wind turbine quality issues persist; grid equipment margins <10%",
    "GE":   "Commercial aerospace orders decline 2+ years; LEAP engine issues",
    "WM":   "FCF margins compress <12% sustained; Stericycle integration fails",
    "ODFL": "LTL market share loss >3pts; pricing discipline breaks industry-wide",
    "TYL":  "Cloud bookings growth <10%; government budget cuts impair pipeline",
    "FICO": "VantageScore gains >30% GSE market share; CFPB regulatory action on pricing",
    "CPRT": "Insurance carriers vertically integrate salvage; total loss rates decline structurally",
    "IDXX": "Vet visit frequency declines sustained; loses reference lab share to Zoetis",
    "VRSN": ".com/.net registry contract not renewed; ICANN policy change",
    "ROP":  "Organic growth <4% for 2+ years; acquisition integration failures",
    "RSG":  "FCF margins compress <12% sustained; loses municipal contracts to WM",
    "JKHY": "Core banking share loss to FIS/Fiserv; credit union consolidation accelerates",
    "PWR":  "Utility CapEx cycle reverses; grid buildout delays >2 years",
}

# 1.7 - Alert Thresholds
SINGLE_DAY_DROP_THRESHOLD = 5.0      # Track 2 trigger: >= 5% drop
NEAR_52_WEEK_LOW_THRESHOLD = 15.0    # Alert if within 15% of 52-week low
BELOW_50_DAY_MA_THRESHOLD = 3.0      # Only alert if >= 3% below 50-day MA
TRACK3_DISTANCE_THRESHOLD = 10.0     # Track 3: within 10% of target

# 1.8 - Track 2 Regime Detection
TRACK2_REGIME_THRESHOLD = 5          # If >= 5 triggers in 10 days, warn
TRACK2_REGIME_WINDOW_DAYS = 10


# =============================================================================
# BLOCK 2: DATA FETCHING
# =============================================================================

def fetch_all_market_data():
    """
    2.1 - Fetch comprehensive market data for all tickers
    Returns dict with price, change, 52-week range, 50-day MA, P/E ratios
    """
    print("\n[1] FETCHING MARKET DATA")
    print("-" * 50)
    
    market_data = {}
    trade_dates = []
    
    for ticker in TICKERS:
        try:
            print(f"    {ticker}...", end=" ")
            stock = yf.Ticker(ticker)
            hist = stock.history(period="400d")
            
            if hist.empty or len(hist) < 2:
                print("NO DATA")
                continue
            
            # 2.1.1 - Basic price data
            today = hist.iloc[-1]
            yesterday = hist.iloc[-2]
            close = today['Close']
            low = today['Low']
            prev_close = yesterday['Close']
            change_pct = ((close - prev_close) / prev_close) * 100
            
            # 2.1.2 - 50-day moving average
            ma_50 = hist['Close'].tail(50).mean() if len(hist) >= 50 else hist['Close'].mean()
            
            # 2.1.3 - 52-week high/low
            window_52w = hist.tail(252) if len(hist) >= 252 else hist
            week_52_low = window_52w['Low'].min()
            week_52_high = window_52w['High'].max()
            
            # 2.1.4 - P/E ratios
            trailing_pe = None
            forward_pe = None
            try:
                info = stock.info
                trailing_pe = info.get('trailingPE')
                forward_pe = info.get('forwardPE')
            except:
                pass
            
            # 2.1.5 - Collect trade date
            ticker_date = hist.index[-1].strftime("%Y-%m-%d")
            trade_dates.append(ticker_date)
            
            market_data[ticker] = {
                'close': close,
                'low': low,
                'prev_close': prev_close,
                'change_pct': change_pct,
                'ma_50': ma_50,
                'week_52_low': week_52_low,
                'week_52_high': week_52_high,
                'trailing_pe': trailing_pe,
                'forward_pe': forward_pe,
            }
            
            print(f"${close:.2f} ({change_pct:+.1f}%)")
            
        except Exception as e:
            print(f"ERROR: {e}")
    
    # 2.1.6 - Determine trade date (mode of all dates)
    if trade_dates:
        from collections import Counter
        date_counts = Counter(trade_dates)
        trade_date = date_counts.most_common(1)[0][0]
        market_data['_trade_date'] = trade_date
        print(f"\n    Trade date: {trade_date}")
    
    print(f"    Fetched: {len([k for k in market_data if not k.startswith('_')])} of {len(TICKERS)} tickers")
    
    return market_data


# =============================================================================
# BLOCK 3: TRACKER UPDATE FUNCTIONS
# =============================================================================

def load_tracker():
    """3.1 - Load the tracker workbook"""
    if not os.path.exists(TRACKER_FILE):
        print(f"    ERROR: Tracker not found: {TRACKER_FILE}")
        return None
    return load_workbook(TRACKER_FILE)


def normalize_date(date_val):
    """3.2 - Normalize date to YYYY-MM-DD string"""
    if date_val is None:
        return ""
    if isinstance(date_val, str):
        return date_val.strip()
    if hasattr(date_val, 'strftime'):
        return date_val.strftime("%Y-%m-%d")
    if isinstance(date_val, (int, float)) and date_val > 40000:
        excel_epoch = datetime(1899, 12, 30)
        return (excel_epoch + timedelta(days=int(date_val))).strftime("%Y-%m-%d")
    return str(date_val)


def update_positions_prices(wb, market_data):
    """3.3 - Update Current_Price column in Positions sheet"""
    print("\n[2] UPDATING POSITION PRICES")
    print("-" * 50)
    
    ws = wb['Positions']
    updated = 0
    
    for row in range(2, ws.max_row + 1):
        ticker = ws.cell(row=row, column=1).value
        if ticker and ticker in market_data:
            price = market_data[ticker]['close']
            ws.cell(row=row, column=5, value=price)
            updated += 1
    
    print(f"    Updated {updated} position prices")
    return updated


def get_cash_balance(wb):
    """3.4 - Get current cash balance from Positions sheet"""
    ws = wb['Positions']
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "CASH":
            return ws.cell(row=row, column=6).value or 0
    return 0


def update_cash_balance(wb, new_balance):
    """3.5 - Update cash balance in Positions sheet"""
    ws = wb['Positions']
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "CASH":
            ws.cell(row=row, column=6, value=new_balance)
            return True
    return False


def update_position(wb, ticker, shares_to_add, price, buy_date):
    """3.6 - Update a position after a fill"""
    ws = wb['Positions']
    
    # Find ticker row
    row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == ticker:
            row = r
            break
    
    if row is None:
        print(f"    ERROR: {ticker} not found in Positions")
        return False
    
    # Calculate new average cost
    current_shares = ws.cell(row=row, column=2).value or 0
    current_avg_cost = ws.cell(row=row, column=3).value or 0
    current_invested = current_shares * current_avg_cost
    
    new_invested = shares_to_add * price
    total_shares = current_shares + shares_to_add
    total_invested = current_invested + new_invested
    new_avg_cost = total_invested / total_shares if total_shares > 0 else 0
    
    ws.cell(row=row, column=2, value=total_shares)
    ws.cell(row=row, column=3, value=new_avg_cost)
    
    # Set first buy date if not set
    if not ws.cell(row=row, column=10).value:
        ws.cell(row=row, column=10, value=buy_date)
    
    return True


def reconcile_pending_orders(wb, market_data):
    """3.7 - Check pending orders against day's low for fills"""
    print("\n[3] RECONCILING PENDING ORDERS")
    print("-" * 50)
    
    ws_pending = wb['Pending_Orders']
    ws_log = wb['Action_Log']
    
    trade_date = market_data.get('_trade_date', datetime.now().strftime("%Y-%m-%d"))
    
    fills = []
    expirations = []
    kept = []
    rows_to_delete = []
    
    cash = get_cash_balance(wb)
    
    for row in range(2, ws_pending.max_row + 1):
        order_date_raw = ws_pending.cell(row=row, column=1).value
        ticker = ws_pending.cell(row=row, column=2).value
        limit_price = ws_pending.cell(row=row, column=3).value
        shares = ws_pending.cell(row=row, column=4).value
        track = ws_pending.cell(row=row, column=5).value
        signal = ws_pending.cell(row=row, column=6).value
        thesis = ws_pending.cell(row=row, column=7).value
        notes = ws_pending.cell(row=row, column=8).value
        status = ws_pending.cell(row=row, column=9).value
        
        if status and status not in ['PENDING', '']:
            continue
        if not ticker or not shares or not limit_price:
            continue
        
        order_date = normalize_date(order_date_raw)
        
        # Skip future-dated orders
        if order_date and trade_date and order_date > trade_date:
            continue
        
        # Check if filled
        if ticker in market_data:
            day_low = market_data[ticker]['low']
            fill_cost = shares * limit_price
            
            if day_low <= limit_price:
                if fill_cost > cash:
                    print(f"    BLOCKED: {ticker} - Insufficient cash")
                    continue
                
                fills.append({
                    'ticker': ticker,
                    'shares': int(shares),
                    'price': limit_price,
                    'track': track,
                    'signal': signal,
                    'thesis': thesis,
                    'notes': notes
                })
                rows_to_delete.append(row)
                cash -= fill_cost
                print(f"    FILLED: {ticker} @ ${limit_price:.2f} (low: ${day_low:.2f})")
            
            elif str(track) in ['1', '2']:
                # Track 1/2 are DAY orders: expire if trade_date >= order_date and not filled
                # This means they had their chance (on order_date) and missed
                if order_date and trade_date and trade_date >= order_date:
                    expirations.append({'ticker': ticker, 'limit': limit_price})
                    rows_to_delete.append(row)
                    print(f"    EXPIRED: {ticker} DAY order @ ${limit_price:.2f} (order date: {order_date})")
                else:
                    # Future-dated DAY order, keep for now
                    kept.append({'ticker': ticker, 'limit': limit_price, 'note': 'future DAY order'})
            else:
                kept.append({'ticker': ticker, 'limit': limit_price})
                print(f"    KEPT: {ticker} GTC @ ${limit_price:.2f}")
    
    # Process fills
    for fill in fills:
        log_row = ws_log.max_row + 1
        ws_log.cell(row=log_row, column=1, value=trade_date)
        ws_log.cell(row=log_row, column=2, value=fill['ticker'])
        ws_log.cell(row=log_row, column=3, value=fill['track'])
        ws_log.cell(row=log_row, column=4, value="BUY")
        ws_log.cell(row=log_row, column=5, value=fill['shares'])
        ws_log.cell(row=log_row, column=6, value=fill['price'])
        ws_log.cell(row=log_row, column=7, value=fill['signal'])
        ws_log.cell(row=log_row, column=8, value=fill['thesis'])
        ws_log.cell(row=log_row, column=9, value=fill['notes'])
        
        update_position(wb, fill['ticker'], fill['shares'], fill['price'], trade_date)
    
    update_cash_balance(wb, cash)
    
    # Delete processed rows
    for row in sorted(rows_to_delete, reverse=True):
        ws_pending.delete_rows(row)
    
    print(f"\n    Fills: {len(fills)}, Expirations: {len(expirations)}, Kept: {len(kept)}")
    print(f"    Cash balance: ${cash:,.2f}")
    
    return fills, expirations, kept


def append_daily_snapshot(wb, market_data):
    """3.8 - Append a row to Benchmark sheet"""
    print("\n[4] APPENDING DAILY SNAPSHOT")
    print("-" * 50)
    
    ws = wb['Benchmark']
    ws_pos = wb['Positions']
    
    trade_date = market_data.get('_trade_date', datetime.now().strftime("%Y-%m-%d"))
    voo_price = market_data.get('VOO', {}).get('close', 0)
    
    # Calculate portfolio value
    total_market_value = 0
    cash = 0
    for row in range(2, ws_pos.max_row + 1):
        label = ws_pos.cell(row=row, column=1).value
        if label == "CASH":
            cash = ws_pos.cell(row=row, column=6).value or 0
        elif label not in ["TOTAL", "PORTFOLIO"]:
            shares = ws_pos.cell(row=row, column=2).value or 0
            price = ws_pos.cell(row=row, column=5).value or 0
            total_market_value += shares * price
    
    portfolio_value = total_market_value + cash
    
    # Calculate returns
    baseline_voo = ws.cell(row=2, column=2).value or voo_price
    baseline_portfolio = ws.cell(row=2, column=3).value or portfolio_value
    
    voo_return = ((voo_price - baseline_voo) / baseline_voo) if baseline_voo else 0
    portfolio_return = ((portfolio_value - baseline_portfolio) / baseline_portfolio) if baseline_portfolio else 0
    alpha = portfolio_return - voo_return
    
    # Append row
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=1, value=trade_date)
    ws.cell(row=new_row, column=2, value=voo_price)
    ws.cell(row=new_row, column=3, value=portfolio_value)
    ws.cell(row=new_row, column=4, value=voo_return)
    ws.cell(row=new_row, column=5, value=portfolio_return)
    ws.cell(row=new_row, column=6, value=alpha)
    
    # Format percentages
    for col in [4, 5, 6]:
        ws.cell(row=new_row, column=col).number_format = '0.0%'
    
    print(f"    Date: {trade_date}")
    print(f"    VOO: ${voo_price:.2f} ({voo_return:+.1%})")
    print(f"    Portfolio: ${portfolio_value:,.2f} ({portfolio_return:+.1%})")
    print(f"    Alpha: {alpha:+.1%}")
    
    return portfolio_value, voo_return, portfolio_return, alpha


# =============================================================================
# BLOCK 4: ALERT DETECTION
# =============================================================================

def load_track2_history():
    """4.1 - Load Track 2 trigger history"""
    if os.path.exists(TRACK2_HISTORY_FILE):
        try:
            with open(TRACK2_HISTORY_FILE, 'r') as f:
                data = json.load(f)
                return [datetime.fromisoformat(d) for d in data]
        except:
            return []
    return []


def save_track2_history(history):
    """4.2 - Save Track 2 trigger history"""
    data = [d.isoformat() for d in history]
    with open(TRACK2_HISTORY_FILE, 'w') as f:
        json.dump(data, f, indent=2)


def check_regime_status():
    """4.3 - Check if Track 2 should be suspended"""
    history = load_track2_history()
    cutoff = datetime.now() - timedelta(days=TRACK2_REGIME_WINDOW_DAYS)
    recent = [d for d in history if d > cutoff]
    
    if len(recent) >= TRACK2_REGIME_THRESHOLD:
        return True, len(recent)
    return False, len(recent)


def detect_alerts(market_data):
    """
    4.4 - Detect all Track 2 and Track 3 alert candidates
    Returns list of alert dictionaries with all relevant data
    """
    print("\n[5] DETECTING ALERTS")
    print("-" * 50)
    
    alerts = []
    track2_triggers = 0
    
    for ticker, data in market_data.items():
        if ticker.startswith('_'):
            continue
        
        info = WATCHLIST.get(ticker, {})
        targets = TARGETS.get(ticker, {})
        
        alert_signals = []
        is_track2 = False
        is_track3 = False
        
        # 4.4.1 - Check for Track 2 (5%+ single-day drop)
        if data['change_pct'] <= -SINGLE_DAY_DROP_THRESHOLD:
            alert_signals.append(f"SINGLE-DAY DROP: {data['change_pct']:.1f}%")
            is_track2 = True
            track2_triggers += 1
        
        # 4.4.2 - Check for Track 3 (within 10% of target)
        target = targets.get('target', 0)
        if target > 0:
            distance_pct = ((data['close'] - target) / target) * 100
            if distance_pct <= TRACK3_DISTANCE_THRESHOLD:
                alert_signals.append(f"NEAR TARGET: {distance_pct:+.1f}% from ${target}")
                is_track3 = True
        else:
            distance_pct = None
        
        # 4.4.3 - Check for near 52-week low
        distance_from_low = ((data['close'] - data['week_52_low']) / data['week_52_low']) * 100
        if distance_from_low <= NEAR_52_WEEK_LOW_THRESHOLD:
            alert_signals.append(f"NEAR 52-WEEK LOW: {distance_from_low:.1f}% above ${data['week_52_low']:.2f}")
        
        # 4.4.4 - Check for below 50-day MA
        if data['close'] < data['ma_50']:
            pct_below = ((data['ma_50'] - data['close']) / data['ma_50']) * 100
            if pct_below >= BELOW_50_DAY_MA_THRESHOLD:
                alert_signals.append(f"BELOW 50-DAY MA: {pct_below:.1f}% below ${data['ma_50']:.2f}")
        
        # 4.4.5 - If any alerts triggered, add to list
        if alert_signals:
            alerts.append({
                'ticker': ticker,
                'company': info.get('name', ticker),
                'strategy': info.get('strategy', 'N/A'),
                'tier': info.get('tier', 'N/A'),
                'price': data['close'],
                'prev_close': data['prev_close'],
                'change_pct': data['change_pct'],
                'ma_50': data['ma_50'],
                'week_52_low': data['week_52_low'],
                'week_52_high': data['week_52_high'],
                'trailing_pe': data['trailing_pe'],
                'forward_pe': data['forward_pe'],
                'target': target,
                'add_target': targets.get('add_target', 0),
                'target_distance_pct': distance_pct,
                'exit_criteria': EXIT_CRITERIA.get(ticker, 'N/A'),
                'signals': alert_signals,
                'is_track2': is_track2,
                'is_track3': is_track3,
            })
            
            track_label = []
            if is_track2:
                track_label.append("T2")
            if is_track3:
                track_label.append("T3")
            print(f"    {ticker}: [{'/'.join(track_label) or 'WATCH'}] {', '.join(alert_signals)}")
    
    # Update Track 2 history
    if track2_triggers > 0:
        history = load_track2_history()
        for _ in range(track2_triggers):
            history.append(datetime.now())
        save_track2_history(history)
    
    # Sort by priority (Track 2 first, then Track 3, then by distance to target)
    alerts.sort(key=lambda x: (
        not x['is_track2'],
        not x['is_track3'],
        x['target_distance_pct'] if x['target_distance_pct'] is not None else 999
    ))
    
    track2_count = sum(1 for a in alerts if a['is_track2'])
    track3_count = sum(1 for a in alerts if a['is_track3'])
    
    print(f"\n    Total alerts: {len(alerts)}")
    print(f"    Track 2 candidates: {track2_count}")
    print(f"    Track 3 candidates: {track3_count}")
    
    return alerts


# =============================================================================
# BLOCK 5: PROMPT GENERATION
# =============================================================================

def get_current_positions(wb):
    """5.1 - Get current positions from tracker"""
    ws = wb['Positions']
    positions = {}
    
    for row in range(2, ws.max_row + 1):
        ticker = ws.cell(row=row, column=1).value
        shares = ws.cell(row=row, column=2).value or 0
        avg_cost = ws.cell(row=row, column=3).value or 0
        
        if ticker and shares > 0 and ticker not in ['CASH', 'TOTAL', 'PORTFOLIO']:
            positions[ticker] = {
                'shares': shares,
                'avg_cost': avg_cost,
            }
    
    return positions


def get_pending_orders(wb):
    """5.2 - Get pending GTC orders from tracker"""
    ws = wb['Pending_Orders']
    orders = {}
    
    for row in range(2, ws.max_row + 1):
        ticker = ws.cell(row=row, column=2).value
        limit_price = ws.cell(row=row, column=3).value
        shares = ws.cell(row=row, column=4).value
        status = ws.cell(row=row, column=9).value
        
        if ticker and status in ['PENDING', '', None]:
            orders[ticker] = {
                'limit': limit_price,
                'shares': shares,
            }
    
    return orders


def build_roundtable_prompt(alerts, market_data, portfolio_stats, positions, pending_orders, regime_status):
    """
    5.3 - Build the complete AI Roundtable prompt
    """
    trade_date = market_data.get('_trade_date', datetime.now().strftime("%Y-%m-%d"))
    regime_suspended, regime_count = regime_status
    portfolio_value, voo_return, portfolio_return, alpha = portfolio_stats
    
    # 5.3.1 - Header
    prompt = f"""
================================================================================
MR. MARKET AI ROUNDTABLE - {trade_date}
================================================================================

You are participating in the "Tantrums & Targets" AI Roundtable. Three analysts
debate each stock to determine: Is Mr. Market overreacting (opportunity) or
correctly repricing long-term earnings power (avoid)?

THE THREE ANALYSTS:
- The Auditor: Rules, valuation, strict methodology adherence
- The Narrator: Macro context, narrative, sentiment
- The Arbiter: Synthesizes conflicts, forces final decision

DEFAULT BIAS: If no consensus, the decision is NONE. Not trading is valid.

================================================================================
PORTFOLIO STATUS
================================================================================
Portfolio Value: ${portfolio_value:,.2f}
VOO Return: {voo_return:+.1%}
Portfolio Return: {portfolio_return:+.1%}
Alpha vs VOO: {alpha:+.1%}
"""

    # 5.3.2 - Current positions
    if positions:
        prompt += "\nCURRENT POSITIONS:\n"
        for ticker, pos in positions.items():
            current_price = market_data.get(ticker, {}).get('close', 0)
            gain_pct = ((current_price - pos['avg_cost']) / pos['avg_cost'] * 100) if pos['avg_cost'] > 0 else 0
            prompt += f"  {ticker}: {pos['shares']} shares @ ${pos['avg_cost']:.2f} (now ${current_price:.2f}, {gain_pct:+.1f}%)\n"
    else:
        prompt += "\nCURRENT POSITIONS: None\n"
    
    # 5.3.3 - Pending GTC orders
    if pending_orders:
        prompt += "\nPENDING GTC ORDERS:\n"
        for ticker, order in pending_orders.items():
            current_price = market_data.get(ticker, {}).get('close', 0)
            distance = ((current_price - order['limit']) / order['limit'] * 100) if order['limit'] > 0 else 0
            prompt += f"  {ticker}: {order['shares']} shares @ ${order['limit']:.2f} (current ${current_price:.2f}, {distance:+.1f}% away)\n"
    
    # 5.3.4 - Regime warning
    if regime_suspended:
        prompt += f"""
================================================================================
!! REGIME WARNING !!
================================================================================
{regime_count} Track 2 triggers in last 10 days (threshold: {TRACK2_REGIME_THRESHOLD})
This suggests a broad market selloff, not stock-specific opportunities.
DEFAULT: Skip Track 2 entries. Focus on Track 3 targets only.
================================================================================
"""
    
    # 5.3.5 - Alert candidates
    if alerts:
        prompt += f"""
================================================================================
TODAY'S CANDIDATES ({len(alerts)} stocks triggered alerts)
================================================================================
"""
        for i, alert in enumerate(alerts, 1):
            track_labels = []
            if alert['is_track2']:
                track_labels.append("TRACK 2")
            if alert['is_track3']:
                track_labels.append("TRACK 3")
            track_str = " + ".join(track_labels) if track_labels else "WATCH"
            
            # P/E info
            pe_parts = []
            if alert['trailing_pe']:
                pe_parts.append(f"T:{alert['trailing_pe']:.1f}")
            if alert['forward_pe']:
                pe_parts.append(f"F:{alert['forward_pe']:.1f}")
            pe_str = " / ".join(pe_parts) if pe_parts else "N/A"
            
            # P/E warning
            pe_warning = ""
            if (alert['trailing_pe'] and alert['trailing_pe'] >= 55) or (alert['forward_pe'] and alert['forward_pe'] >= 35):
                pe_warning = " ** P/E HIGH: Extra skepticism warranted **"
            
            # Position status
            position_status = ""
            if alert['ticker'] in positions:
                pos = positions[alert['ticker']]
                position_status = f"\n    OWNED: {pos['shares']} shares @ ${pos['avg_cost']:.2f}"
            if alert['ticker'] in pending_orders:
                order = pending_orders[alert['ticker']]
                position_status += f"\n    GTC ORDER: {order['shares']} shares @ ${order['limit']:.2f}"
            
            prompt += f"""
--------------------------------------------------------------------------------
CANDIDATE {i}: {alert['ticker']} ({alert['company']}) [{track_str}]
--------------------------------------------------------------------------------
Strategy: {alert['strategy']} | Tier: {alert['tier']}

SIGNALS:
"""
            for signal in alert['signals']:
                prompt += f"  - {signal}\n"
            
            prompt += f"""
SNAPSHOT:
  - Price: ${alert['price']:.2f} (prev close: ${alert['prev_close']:.2f}, change: {alert['change_pct']:+.1f}%)
  - 52-week range: ${alert['week_52_low']:.2f} - ${alert['week_52_high']:.2f}
  - 50-day MA: ${alert['ma_50']:.2f}
  - P/E: {pe_str}{pe_warning}
  - Target: ${alert['target']} (START) / ${alert['add_target']} (ADD)
  - Distance to target: {alert['target_distance_pct']:+.1f}%{position_status}

EXIT CRITERIA (what breaks the thesis):
  {alert['exit_criteria']}
"""
            # Track 2 specific
            if alert['is_track2']:
                prompt += f"""
TRACK 2 ENTRY RULES:
  - Entry limit: ${alert['prev_close']:.2f} (prior close, exact)
  - Order type: DAY order - if not filled by close, mark as MISSED
  - Reassessment: ~2 weeks after entry (TRIM if +4%, CONVERT or CUT)
"""
            
            # Track 3 specific
            if alert['is_track3']:
                prompt += f"""
TRACK 3 ENTRY RULES:
  - GTC limit order at ${alert['target']:.2f}
  - If already have GTC working, maintain it
  - ADD target at ${alert['add_target']:.2f} for second tranche
"""

    else:
        prompt += """
================================================================================
NO ALERTS TRIGGERED
================================================================================
Mr. Market is calm today. No Track 2 or Track 3 candidates.
Review any existing positions and pending orders only.
"""
    
    # 5.3.6 - Deliverables template
    prompt += """
================================================================================
DELIVERABLES FOR EACH CANDIDATE
================================================================================
For each candidate above, provide:

1) CATALYST: Most likely reason for the recent move (with dates if known)

2) VALUATION: Forward P/E, trailing P/E, comparison to 5-year average
   - Cite sources or mark as "uncited"

3) RISK TYPE: cyclical / execution / structural / legal-reg / tech narrative / balance sheet

4) BEAR CASE: Steelman the strongest bearish argument

5) BULL REBUTTAL: Steelman the strongest bullish counter

6) IMPAIRMENT TEST: What must be true for long-term earnings power to be permanently damaged?

7) CHECKPOINTS: 3 falsifiable items to monitor over next 1-2 quarters

8) NEXT READ: Specific 10-K sections, earnings call topics, or filings to review

9) DECISION: For each analyst, state one of:
   - IGNORE (not worth attention)
   - WATCH (interesting but not actionable)
   - START SMALL (initiate position at target)
   - ADD (increase existing position)
   - HOLD (maintain current position, no action)
   
   Include:
   - Confidence level (%)
   - Upgrade/downgrade triggers
   - Suggested position size if applicable

================================================================================
FINAL SYNTHESIS (The Arbiter)
================================================================================
After all three analysts have weighed in on each candidate:

1) Identify where analysts AGREE vs DISAGREE
2) Resolve conflicts with explicit reasoning
3) State FINAL DECISION for each candidate
4) List any CALENDAR FLAGS (earnings dates, catalysts to watch)

Rules:
- Be explicit about uncertainty
- Cite sources; label uncited statements as such
- Default to NONE if no clear consensus

================================================================================
DECISION OUTPUT FORMAT (Required for ingestion)
================================================================================
For each actionable decision (BUY/ADD), output a block in this EXACT format:

DECISION:
Date: {today's date, YYYY-MM-DD}
Action: BUY
Ticker: {TICKER}
Limit: {limit price, numbers only}
Shares: {number of shares}
Track: {1, 2, or 3}
Signal: {brief signal description}
Thesis: {Intact/Weakening/Broken}
Notes: {one-line rationale}

Example:
DECISION:
Date: 2026-01-26
Action: BUY
Ticker: ROP
Limit: 400.00
Shares: 10
Track: 3
Signal: Target Hit
Thesis: Intact
Notes: Quality compounder at valuation floor, organic growth intact

For NONE/HOLD/WATCH decisions, still output a block but with Action: NONE
(these will be logged but no order created)

DECISION:
Date: 2026-01-26
Action: NONE
Ticker: MSFT
Limit: 0
Shares: 0
Track: 3
Signal: N/A
Thesis: Intact
Notes: Already own 10 shares, wait for ADD target at $445
================================================================================
"""
    
    return prompt.strip()


def save_prompt_to_file(prompt, trade_date):
    """5.4 - Save prompt to daily file"""
    filename = f"{trade_date.replace('-', '_')}_roundtable_prompt.txt"
    filepath = os.path.join(PROMPTS_DIR, filename)
    
    with open(filepath, 'w') as f:
        f.write(prompt)
    
    return filepath


# =============================================================================
# BLOCK 6: MAIN EXECUTION
# =============================================================================

def main():
    """6.1 - Main execution"""
    
    # 6.1.0 - Parse command line arguments
    parser = argparse.ArgumentParser(
        description='Mr. Market Roundtable - Combined daily script',
        epilog='Example: python mr_market_roundtable.py --decision decisions/2026_01_26_decision.txt'
    )
    parser.add_argument(
        '--decision', '-d',
        help='Path to decision file to ingest into Pending_Orders'
    )
    parser.add_argument(
        '--ingest-only',
        action='store_true',
        help='Only ingest decisions, skip market data fetch and alerts'
    )
    args = parser.parse_args()
    
    # 6.1.1 - Load tracker (needed for both modes)
    wb = load_tracker()
    if wb is None:
        print("ERROR: Could not load tracker. Exiting.")
        return
    
    # 6.1.2 - If ingest-only mode, just process decisions and exit
    if args.ingest_only:
        if not args.decision:
            print("ERROR: --ingest-only requires --decision file")
            return
        print("\n[INGEST-ONLY MODE]")
        added, skipped, rejected = ingest_decisions(wb, args.decision)
        wb.save(TRACKER_FILE)
        print(f"\n    Saved: {TRACKER_FILE}")
        print(f"    Added: {added}, Skipped: {skipped}, Rejected: {rejected}")
        return
    
    # 6.1.3 - Fetch market data
    market_data = fetch_all_market_data()
    if not market_data:
        print("ERROR: No market data fetched. Exiting.")
        return
    
    trade_date = market_data.get('_trade_date', datetime.now().strftime("%Y-%m-%d"))
    
    # 6.1.4 - Update tracker state
    update_positions_prices(wb, market_data)
    fills, expirations, kept = reconcile_pending_orders(wb, market_data)
    portfolio_stats = append_daily_snapshot(wb, market_data)
    
    # 6.1.5 - Detect alerts
    alerts = detect_alerts(market_data)
    
    # 6.1.6 - Ingest decisions if provided (AFTER state is updated)
    decisions_added = 0
    if args.decision:
        print("\n[7] INGESTING DECISIONS")
        print("-" * 50)
        added, skipped, rejected = ingest_decisions(wb, args.decision)
        decisions_added = added
    
    # 6.1.7 - Save tracker
    wb.save(TRACKER_FILE)
    print(f"\n    Saved: {TRACKER_FILE}")
    
    # 6.1.8 - Get current state for prompt (after all updates)
    wb = load_tracker()  # Reload to get fresh state
    positions = get_current_positions(wb)
    pending_orders = get_pending_orders(wb)
    regime_status = check_regime_status()
    
    # 6.1.9 - Build and save prompt
    print("\n[8] GENERATING ROUNDTABLE PROMPT")
    print("-" * 50)
    
    prompt = build_roundtable_prompt(
        alerts, 
        market_data, 
        portfolio_stats, 
        positions, 
        pending_orders, 
        regime_status
    )
    
    filepath = save_prompt_to_file(prompt, trade_date)
    print(f"    Saved prompt: {filepath}")
    
    # 6.1.10 - Summary
    print("\n" + "=" * 70)
    print("ROUNDTABLE COMPLETE")
    print("=" * 70)
    print(f"  Trade date: {trade_date}")
    print(f"  Fills today: {len(fills)}")
    print(f"  Alerts triggered: {len(alerts)}")
    print(f"  Track 2 candidates: {sum(1 for a in alerts if a['is_track2'])}")
    print(f"  Track 3 candidates: {sum(1 for a in alerts if a['is_track3'])}")
    if args.decision:
        print(f"  Decisions ingested: {decisions_added}")
    print(f"\n  PROMPT FILE: {filepath}")
    print(f"\n  Attach this file to ChatGPT, Gemini, and Claude for today's roundtable.")
    print("=" * 70)


# =============================================================================
# BLOCK 7: DECISION INGESTION
# =============================================================================

def parse_decision_blocks(text):
    """
    7.1 - Parse DECISION blocks from text file
    
    Expected format:
    DECISION:
    Date: 2026-01-26
    Action: BUY
    Ticker: ROP
    Limit: 400.00
    Shares: 10
    Track: 3
    Signal: Target Hit
    Thesis: Intact
    Notes: Quality compounder at valuation floor
    """
    decisions = []
    
    # 7.1.0 - Normalize line endings (handles Windows \r\n and old Mac \r)
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    
    # 7.1.1 - Find all DECISION blocks
    pattern = r'DECISION(?:\s*\d*)?:\s*\n((?:.*?\n)*?)(?=DECISION|\Z)'
    matches = re.findall(pattern, text, re.IGNORECASE)
    
    for match in matches:
        decision = {}
        lines = match.strip().split('\n')
        
        for line in lines:
            if ':' in line:
                key, value = line.split(':', 1)
                key = key.strip().lower()
                value = value.strip()
                
                if key == 'date':
                    decision['date'] = value
                elif key == 'action':
                    decision['action'] = value.upper()
                elif key == 'ticker':
                    decision['ticker'] = value.upper()
                elif key == 'limit':
                    try:
                        decision['limit'] = float(value.replace('$', '').replace(',', ''))
                    except:
                        decision['limit'] = 0
                elif key == 'shares':
                    try:
                        decision['shares'] = int(float(value))
                    except:
                        decision['shares'] = 0
                elif key == 'track':
                    decision['track'] = str(value).strip()
                elif key == 'signal':
                    decision['signal'] = value
                elif key == 'thesis':
                    decision['thesis'] = value
                elif key == 'notes':
                    decision['notes'] = value
        
        # 7.1.2 - Only add if we have minimum required fields
        if decision.get('action') and decision.get('ticker'):
            decisions.append(decision)
    
    return decisions


def get_pending_order_key(order):
    """
    7.2 - Generate unique key for idempotency check
    """
    date_str = normalize_date(order.get('date', ''))
    ticker_str = str(order.get('ticker', '')).upper().strip()
    track_str = str(order.get('track', '')).strip()
    limit_val = order.get('limit', 0)
    limit_str = f"{float(limit_val):.2f}" if limit_val else "0.00"
    shares_val = order.get('shares', 0)
    shares_str = str(int(float(shares_val))) if shares_val else "0"
    
    return f"{date_str}|{ticker_str}|{track_str}|{limit_str}|{shares_str}"


def order_exists(wb, order):
    """
    7.3 - Check if order already exists in Pending_Orders or Action_Log
    """
    key = get_pending_order_key(order)
    
    # 7.3.1 - Check Pending_Orders
    ws = wb['Pending_Orders']
    for row in range(2, ws.max_row + 1):
        existing = {
            'date': ws.cell(row=row, column=1).value,
            'ticker': ws.cell(row=row, column=2).value,
            'track': ws.cell(row=row, column=5).value,
            'limit': ws.cell(row=row, column=3).value,
            'shares': ws.cell(row=row, column=4).value
        }
        if get_pending_order_key(existing) == key:
            return True
    
    # 7.3.2 - Check Action_Log (already executed)
    ws = wb['Action_Log']
    for row in range(2, ws.max_row + 1):
        existing = {
            'date': ws.cell(row=row, column=1).value,
            'ticker': ws.cell(row=row, column=2).value,
            'track': ws.cell(row=row, column=3).value,
            'limit': ws.cell(row=row, column=6).value,
            'shares': ws.cell(row=row, column=5).value
        }
        if get_pending_order_key(existing) == key:
            return True
    
    return False


def ingest_decisions(wb, decision_file):
    """
    7.4 - Read DECISION blocks from file and add to Pending_Orders
    
    Returns: (added_count, skipped_count, rejected_count)
    """
    # 7.4.1 - Check file exists
    if not decision_file or not os.path.exists(decision_file):
        print(f"    ERROR: Decision file not found: {decision_file}")
        return 0, 0, 0
    
    print(f"    Reading: {decision_file}")
    
    # 7.4.2 - Read and parse file
    with open(decision_file, 'r') as f:
        text = f.read()
    
    decisions = parse_decision_blocks(text)
    print(f"    Found {len(decisions)} decision block(s)")
    
    ws = wb['Pending_Orders']
    added = 0
    skipped = 0
    rejected = 0
    
    # 7.4.3 - Get current cash for validation
    cash = get_cash_balance(wb)
    
    for decision in decisions:
        ticker = decision.get('ticker', '')
        action = decision.get('action', '')
        shares = decision.get('shares', 0)
        limit = decision.get('limit', 0)
        track = decision.get('track', '')
        
        # 7.4.4 - Handle non-actionable decisions (no order needed)
        if action in ['NONE', 'HOLD', 'WATCH', 'IGNORE']:
            print(f"    SKIP: {ticker} - Action is {action} (no order)")
            skipped += 1
            continue
        
        # 7.4.5 - Validate action is BUY or ADD (the only actionable types we support)
        if action not in ['BUY', 'ADD']:
            print(f"    REJECTED: {ticker} - Action must be BUY or ADD (got: {action})")
            rejected += 1
            continue
        
        # 7.4.6 - Validate ticker is in watchlist
        if ticker not in TICKERS:
            print(f"    REJECTED: {ticker} - Not in 25-stock watchlist")
            rejected += 1
            continue
        
        # 7.4.7 - Validate shares >= 1
        if not shares or shares < 1:
            print(f"    REJECTED: {ticker} - Shares must be >= 1 (got: {shares})")
            rejected += 1
            continue
        
        # 7.4.8 - Validate limit > 0
        if not limit or limit <= 0:
            print(f"    REJECTED: {ticker} - Limit must be > 0 (got: {limit})")
            rejected += 1
            continue
        
        # 7.4.9 - Validate track in {1, 2, 3}
        if track not in ['1', '2', '3']:
            print(f"    REJECTED: {ticker} - Track must be 1, 2, or 3 (got: {track})")
            rejected += 1
            continue
        
        # 7.4.10 - Idempotency check
        if order_exists(wb, decision):
            print(f"    SKIP: {ticker} - Order already exists (idempotency)")
            skipped += 1
            continue
        
        # 7.4.11 - Cash check
        cost = shares * limit
        if cost > cash:
            print(f"    REJECTED: {ticker} - Insufficient cash (need ${cost:.2f}, have ${cash:.2f})")
            rejected += 1
            continue
        
        # 7.4.12 - Decrement cash for next decision (prevent over-subscription)
        cash -= cost
        
        # 7.4.13 - Add to Pending_Orders
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1, value=decision.get('date', datetime.now().strftime('%Y-%m-%d')))
        ws.cell(row=new_row, column=2, value=ticker)
        ws.cell(row=new_row, column=3, value=limit)
        ws.cell(row=new_row, column=4, value=int(shares))
        ws.cell(row=new_row, column=5, value=track)
        ws.cell(row=new_row, column=6, value=decision.get('signal', ''))
        ws.cell(row=new_row, column=7, value=decision.get('thesis', ''))
        ws.cell(row=new_row, column=8, value=decision.get('notes', ''))
        ws.cell(row=new_row, column=9, value='PENDING')
        
        print(f"    ADDED: {ticker} @ ${limit:.2f} x {int(shares)} shares (Track {track})")
        added += 1
    
    print(f"\n    Summary: Added {added}, Skipped {skipped}, Rejected {rejected}")
    return added, skipped, rejected


if __name__ == "__main__":
    main()